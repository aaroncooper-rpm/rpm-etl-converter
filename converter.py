"""
RPM Living · Yardi ETL Conversion Engine
All property-specific mappings are loaded dynamically from the RPM Takeover Guide.
No hardcoded property data — works for any RPM Living property.
"""
import pandas as pd
import numpy as np
import re
import zipfile
import os
import shutil
from datetime import datetime, date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter



def load_former_records(base):
    """
    Load former residents who must still appear in Yardi:
      - Eviction proceedings started (any balance)    → status 10
      - Former resident with non-zero Ledger Balance  → status 5
    Returns a dict: resh_id → tenant record (same shape as build_tenant_base output)
    """
    try:
        ld = pd.read_excel(_find_file(base, "Lease Details .xlsx"), header=8)
    except Exception:
        return {}

    ld["resh_id"] = ld["Household ID/ Resh ID"].apply(lambda x: safe_int(x, 0))
    ld = ld[ld["resh_id"] > 0]

    former = ld[ld["Occupancy status"].str.contains("Former", case=False, na=False)]
    former = former.drop_duplicates("resh_id", keep="last")

    eviction_ids = set(
        ld[ld["Eviction proceedings started"] == "Yes"]["resh_id"].dropna().astype(int)
    )
    balance_ids = set(
        former[former["Ledger Balance"].fillna(0) != 0]["resh_id"].astype(int)
    )

    include_ids = eviction_ids | balance_ids
    subset = former[former["resh_id"].isin(include_ids)].copy()

    tenants = {}
    for _, row in subset.iterrows():
        resh_id = int(row["resh_id"])
        if resh_id in tenants:
            continue

        name_raw = str(row.get("Household name", "")) if pd.notna(row.get("Household name")) else ""
        last, first, middle = parse_name(name_raw)

        is_eviction = resh_id in eviction_ids
        status = 10 if is_eviction else 5

        # Parse unit code from "Unit #" column (stored as '0122 style)
        unit_raw = str(row.get("Unit #", "")) if pd.notna(row.get("Unit #")) else ""
        unit_code = clean_unit(unit_raw)

        phone1  = clean_phone(row.get("Cell Phone"))
        phone2  = clean_phone(row.get("Home Phone"))
        phone3  = clean_phone(row.get("Work Phone"))
        move_in = fmt_date(row.get("Move-in date"))
        move_out= fmt_date(row.get("Moved out date"))
        notice  = fmt_date(row.get("Notice given date"))
        l_from  = fmt_date(row.get("Lease start date"))
        l_to    = fmt_date(row.get("Lease end date"))
        l_sign  = fmt_date(row.get("Lease signed date"))
        l_term  = extract_term(row.get("Lease term")) or 12
        balance = float(row.get("Ledger Balance") or 0)
        l_rent  = safe_int(row.get("Lease Rent"))
        deposit = safe_int(row.get("Required deposit"))

        # Try to parse address from forwarding/billing
        address1 = city = state = zipcode = None
        for af in ("FAS Forwarding address", "Billing address"):
            addr = row.get(af, "")
            if pd.notna(addr) and str(addr).strip():
                import re as _re
                m = _re.match(r"(.+?)\s*,\s*(.+?)\s*,\s*([A-Z]{2})\s+([\d\-]+)", str(addr))
                if m:
                    address1 = m.group(1).strip(); city = m.group(2).strip()
                    state    = m.group(3).strip(); zipcode = m.group(4).strip()
                    break

        tenants[resh_id] = {
            "resh_id":     resh_id,
            "tenant_code": make_tenant_code(resh_id),
            "unit_code":   unit_code,
            "status":      status,
            "last_name":   last,
            "first_name":  first,
            "middle_name": middle or None,
            "dob":         None,
            "ssn":         None,
            "move_in":     move_in,
            "move_out":    move_out,
            "notice_date": notice,
            "lease_from":  l_from,
            "lease_to":    l_to,
            "lease_sign":  l_sign,
            "phone1": phone1, "phone2": phone2, "phone3": phone3, "phone4": None,
            "email":       None,
            "address1":    address1, "address2": None,
            "city": city, "state": state, "zipcode": zipcode,
            "rent":    l_rent,
            "deposit": deposit,
            "lease_term": l_term,
            "floorplan":  "",
        }
    return tenants


# ── Report-title fingerprints ────────────────────────────────────────────────
# Each tuple is (keywords_that_must_all_appear_in_first_12_rows, logical_name)
# Matched case-insensitively against the first 12 rows × first 4 columns.
_REPORT_FINGERPRINTS = {
    # Fingerprints use 2+ unique keywords that together appear only in that report.
    # More-specific entries are first; "ALL RESIDENTS" is last as a broad fallback.
    "Final Rent Roll Detail with Lease Charges.xls": ["RENT ROLL DETAIL"],
    "Resident Birthdays.xls":                        ["RESIDENT BIRTHDAYS"],
    "Final All Unit.xls":                            ["ALL UNITS", "Amt/SQFT"],
    "Renters Insurance Status Report.xlsx":          ["Renters Insurance Status Report"],
    "Prospect Contact Level Details .xlsx":          ["Prospect Contact Level Details", "Guest Card ID"],
    "Contract Level Detail.xlsx":                    ["Contact Level Details", "Emergency Contact"],
    "Final Unit Setup View.xlsx":                    ["Unit Setup View"],
    "Lease Details .xlsx":                           ["Lease Details", "Eviction proceedings"],
    "Rentable and Assignable Items Status.xls":      ["Item Type - Rentable"],
    "All Residents.xls":                             ["ALL RESIDENTS"],
}

# Cache: base_dir → {logical_name: resolved_path}
_FILE_CACHE: dict[str, dict[str, str]] = {}


def _scan_report_titles(base: str) -> dict[str, str]:
    """Scan all Excel files in base, match by report-title fingerprint."""
    import xlrd as _xlrd
    from openpyxl import load_workbook as _lwb

    result: dict[str, str] = {}

    def _cell_texts(path: str) -> list[str]:
        """Return all non-empty string cell values from first 12 rows (all cols)."""
        texts = []
        try:
            if path.endswith(".xls"):
                wb = _xlrd.open_workbook(path)
                ws = wb.sheet_by_index(0)
                for r in range(min(12, ws.nrows)):
                    for c in range(ws.ncols):
                        v = str(ws.cell_value(r, c)).strip()
                        if v: texts.append(v)
                wb.release_resources()
            else:
                wb = _lwb(path, read_only=True, data_only=True)
                ws = wb.active
                for i, row in enumerate(ws.iter_rows(max_row=12, values_only=True)):
                    if i >= 12: break
                    for v in row:
                        if v: texts.append(str(v).strip())
                wb.close()
        except Exception:
            pass
        return texts

    files = [f for f in os.listdir(base)
             if f.lower().endswith((".xls", ".xlsx")) and not f.startswith("~")]

    # Special case: Resident Birthdays has no title row — match by "Birthday" header
    # We'll handle it via fingerprint keywords that appear in its header rows

    for fname in files:
        path = os.path.join(base, fname)
        texts = _cell_texts(path)
        all_text = " ".join(texts).lower()

        for logical_name, keywords in _REPORT_FINGERPRINTS.items():
            if logical_name in result:
                continue   # already matched
            # All keywords must appear (case-insensitive) in the combined text
            if all(kw.lower() in all_text for kw in keywords):
                result[logical_name] = path
                break

    return result


def _find_file(base: str, name: str) -> str:
    """
    Locate a source file by its report-title content rather than filename.
    Falls back to filename-based lookup if the scanner cannot find a match.
    """
    # Normalise base
    if not base.endswith(os.sep):
        base = base + os.sep

    if base not in _FILE_CACHE:
        _FILE_CACHE[base] = _scan_report_titles(base)

    resolved = _FILE_CACHE[base].get(name)
    if resolved and os.path.exists(resolved):
        return resolved

    # Fallback: try exact filename, then swap extension
    p = base + name
    if os.path.exists(p):
        return p
    if name.endswith(".xls") and os.path.exists(base + name + "x"):
        return base + name + "x"
    if name.endswith(".xlsx") and os.path.exists(base + name[:-1]):
        return base + name[:-1]
    return p  # will produce a clear FileNotFoundError downstream


# ─────────────────────────── TAKEOVER GUIDE PARSER ───────────────────────────

def load_takeover_guide(path):
    """
    Parse the RPM Living Takeover Guide Excel file and return a mappings dict.
    Returns:
        {
          "property_code": str,      # Yardi numeric code (user-supplied or from guide)
          "prop_prefix":   str,      # e.g. "RGB"
          "prop_name":     str,
          "address":       str,
          "city":          str,
          "state":         str,
          "zipcode":       str,
          "unit_type_map": {prior_code: {yardi_code, beds, baths, sqft, rent}},
          "amenity_map":   {prior_name: (rpm_desc, rpm_code, amount)},
          "rentable_types":[{prior_name, type_name, rent, charge_code}],
        }
    """
    wb = load_workbook(path, read_only=True, data_only=True)

    mappings = {
        "property_code": "00000",
        "prop_prefix":   "PROP",
        "prop_name":     "Unknown Property",
        "address":       "",
        "city":          "",
        "state":         "",
        "zipcode":       "",
        "unit_type_map": {},
        "amenity_map":   {},
        "rentable_types": [],
    }

    # ── Property Info ──────────────────────────────────────────────────────
    if "Property Info" in wb.sheetnames:
        ws = wb["Property Info"]
        rows = list(ws.iter_rows(max_row=25, values_only=True))
        for r in rows:
            if r[0] and str(r[0]).strip() == "Property Name:":
                mappings["prop_name"] = str(r[2] or "").strip()
            if r[0] and str(r[0]).strip() == "Address:":
                mappings["address"] = str(r[2] or "").strip()
                # Try to parse city/state/zip from next row
            if r[2] and re.match(r".+,\s*[A-Z]{2}\s+\d{5}", str(r[2])):
                m = re.match(r"(.+),\s*([A-Z]{2})\s+(\d{5})", str(r[2]).strip())
                if m:
                    mappings["city"]    = m.group(1).strip()
                    mappings["state"]   = m.group(2).strip()
                    mappings["zipcode"] = m.group(3).strip()

    # ── Unit Type sheet ────────────────────────────────────────────────────
    if "Unit Type" in wb.sheetnames:
        ws = wb["Unit Type"]
        rows = list(ws.iter_rows(max_row=60, values_only=True))
        prop_prefix_found = False
        for row_idx, r in enumerate(rows):
            if not r[0] or str(r[0]).strip() in ("Prior Unit Type", ""):
                continue
            prior   = str(r[0]).strip()
            yardi   = str(r[1] or "").strip()   # full yardi code e.g. RGBA1
            rpm     = str(r[2] or "").strip()   # short code e.g. A1
            desc    = str(r[4] or "").strip() if r[4] else ""
            beds    = int(r[5]) if r[5] is not None and str(r[5]).replace(".0","").isdigit() else 0
            baths   = float(r[6]) if r[6] is not None else 0
            sqft    = int(r[7]) if r[7] is not None and str(r[7]).replace(".0","").isdigit() else 0
            rent    = int(r[8]) if r[8] is not None and str(r[8]).replace(".0","").isdigit() else 0

            # Col D (index 3) on data rows contains the property prefix
            if not prop_prefix_found and r[3] and str(r[3]).strip() not in ("Property Code", ""):
                mappings["prop_prefix"] = str(r[3]).strip()
                prop_prefix_found = True
            # Fallback: extract prefix from full yardi code (e.g. RGBA1 → RGB)
            if not prop_prefix_found and yardi:
                extracted = re.match(r"([A-Z]+)\d", yardi)
                if extracted:
                    mappings["prop_prefix"] = extracted.group(1)
                    prop_prefix_found = True

            if prior and yardi and beds:
                mappings["unit_type_map"][prior] = {
                    "yardi_code": yardi,
                    "beds":  beds,
                    "baths": baths,
                    "sqft":  sqft,
                    "rent":  rent,
                    "desc":  desc,
                }

    # ── Property Amenities ─────────────────────────────────────────────────
    if "Property Amenities" in wb.sheetnames:
        ws = wb["Property Amenities"]
        for r in ws.iter_rows(max_row=80, values_only=True):
            if not r[0] or str(r[0]).strip() in ("Prior Amentity Name", "Prior Amenity Name", ""):
                continue
            prior    = str(r[0]).strip()
            rpm_desc = str(r[1] or r[0]).strip()
            rpm_code = str(r[2] or r[0]).strip() if r[2] else str(r[1] or r[0]).strip()
            try:
                amount = float(r[3]) if r[3] is not None else 0.0
            except (ValueError, TypeError):
                amount = 0.0
            if prior:
                # Sanitise code: no spaces, max 20 chars
                code_clean = re.sub(r"[^A-Za-z0-9_\-]", "", rpm_code)[:20] or re.sub(r"[^A-Za-z0-9_\-]", "", prior)[:20]
                mappings["amenity_map"][prior] = (rpm_desc, code_clean, amount)

    # ── Rentable Item Types ────────────────────────────────────────────────
    if "Rentable Item Types" in wb.sheetnames:
        ws = wb["Rentable Item Types"]
        rows = list(ws.iter_rows(max_row=30, values_only=True))
        for r in rows:
            if not r[0] or str(r[0]).strip() in ("Current Rentable Item", ""):
                continue
            prior_name = str(r[0]).strip()
            type_name  = str(r[1] or r[0]).strip()
            try:
                rent = float(r[2]) if r[2] is not None else 0.0
            except (ValueError, TypeError):
                rent = 0.0
            # Derive charge code: lowercase, no spaces
            charge_code = re.sub(r"\s+", "", type_name.lower())[:8] + "msc"
            mappings["rentable_types"].append({
                "prior_name":  prior_name,
                "type_name":   type_name,
                "rent":        rent,
                "charge_code": charge_code,
            })

    wb.close()
    return mappings


# ─────────────────────────── HELPERS ─────────────────────────────────────────

def clean_unit(u):
    if pd.isna(u): return None
    s = re.sub(r"['\s#]", "", str(u))
    m = re.match(r"(\d+)", s)
    return m.group(1).zfill(4) if m else None

def make_tenant_code(resh_id):
    return f"t{int(resh_id):07d}"

def parse_name(full_name):
    if not full_name or pd.isna(full_name): return ("", "", "")
    name = str(full_name).split(";")[0].strip()
    if "," in name:
        last, rest = name.split(",", 1)
        parts = rest.strip().split()
        return (last.strip(), parts[0] if parts else "", " ".join(parts[1:]) if len(parts) > 1 else "")
    parts = name.split()
    return (parts[-1] if len(parts) > 1 else parts[0],
            parts[0] if len(parts) > 1 else "",
            " ".join(parts[1:-1]) if len(parts) > 2 else "")

def clean_phone(ph):
    if pd.isna(ph) or str(ph).strip() in ("", "nan"): return None
    digits = re.sub(r"[^\d]", "", str(ph))
    if digits.startswith("1") and len(digits) == 11: digits = digits[1:]
    return digits if 7 <= len(digits) <= 10 else None

def fmt_date(d):
    if d is None or (isinstance(d, float) and np.isnan(d)): return None
    try:
        if isinstance(d, (datetime, date)): return d.strftime("%Y-%m-%d")
        return pd.to_datetime(d).strftime("%Y-%m-%d")
    except: return None

def safe_int(v, default=0):
    if v is None or (isinstance(v, float) and np.isnan(v)): return default
    try: return int(float(str(v).replace(",", "").replace("$", "")))
    except: return default

def extract_term(s):
    if pd.isna(s): return None
    m = re.search(r"(\d+)", str(s))
    return int(m.group(1)) if m else None

def write_etl_xlsx(table_name, columns, rows_data, filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = table_name[:31]
    hdr_fill = PatternFill("solid", start_color="1F4E79")
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
    data_font = Font(name="Calibri", size=9)
    ws.cell(1, 1).value = table_name
    ws.cell(1, 1).font  = Font(name="Calibri", bold=True, size=10)
    for ci, col in enumerate(columns, 1):
        c = ws.cell(2, ci, col)
        c.font  = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for ri, row in enumerate(rows_data, 3):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, val).font = data_font
    for ci, col in enumerate(columns, 1):
        mx = len(str(col))
        for row in rows_data:
            v = row[ci - 1]
            if v is not None: mx = max(mx, len(str(v)))
        ws.column_dimensions[get_column_letter(ci)].width = min(mx + 2, 35)
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"
    wb.save(filepath)


# ─────────────────────────── CHARGE CODE MAP ─────────────────────────────────
# Fixed mapping from OneSite Rent Roll column names → Yardi transaction codes.
# These are standard OneSite column names, not property-specific.
CHARGE_CODE_MAP = {
    "RENT":     "rent",
    "INTERNET": "internet",
    "Pet Rent": "petrent",
    "PREMP":    "premp",
    "GARAGE":   "garagmsc",
    "CABLE":    "cable",
    "PETFEE":   "petfee",
    "RINSUR":   "rinsur",
    "MODEL":    "model",
    "EMPLCRED": "emplcred",
    "CONC/MGR": "concmgr",
}

STATUS_MAP = {
    "Occupied":        0,
    "Pending renewal": 0,
    "Occupied-NTV":    4,
    "Occupied-NTVL":   4,
    "Applicant":       6,
    "Vacant-Leased":   6,   # pre-leased vacant unit - future resident
    # Eviction (10) and Former-with-balance (5) are loaded separately from Lease Details
}

# ─────────────────────────── DATA LOADERS ────────────────────────────────────

def load_rent_roll(base):
    df = pd.read_excel(_find_file(base, "Final Rent Roll Detail with Lease Charges.xls"),
                       engine=("xlrd" if _find_file(base, "Final Rent Roll Detail with Lease Charges.xls").endswith(".xls") else None), header=5)
    df = df[df["Bldg/Unit"].notna()]
    df = df[df["Unit/Lease Status"].isin(STATUS_MAP.keys())]
    df["unit_code"] = df["Bldg/Unit"].apply(clean_unit)
    df["resh_id"]   = df["Resh ID"].apply(lambda x: safe_int(x, 0))
    return df[df["resh_id"] > 0]

def load_lease_details(base):
    df = pd.read_excel(_find_file(base, "Lease Details .xlsx"), header=8)
    df["resh_id"] = df["Household ID/ Resh ID"].apply(lambda x: safe_int(x, 0))
    df = df[df["resh_id"] > 0]
    return df.drop_duplicates("resh_id", keep="last").set_index("resh_id")

def load_contract_details(base):
    df = pd.read_excel(_find_file(base, "Contract Level Detail.xlsx"), header=8)
    df["unit_code"] = df["Unit #"].apply(clean_unit)
    curr = df[df["Lease-level Occupancy Status"].str.contains("Current", case=False, na=False)]
    by_unit = {}
    for _, row in curr.iterrows():
        uc = row["unit_code"]
        if uc: by_unit.setdefault(uc, []).append(row)
    return by_unit

def load_all_residents(base):
    df = pd.read_excel(_find_file(base, "All Residents.xls"),
                       engine=("xlrd" if _find_file(base, "All Residents.xls").endswith(".xls") else None), header=5)
    df = df[df["Bldg/Unit"].notna()]
    df = df[df["Bldg/Unit"].apply(lambda x: bool(re.match(r"^\d", str(x))))]
    df["unit_code"] = df["Bldg/Unit"].apply(clean_unit)
    return df

def load_unit_setup(base):
    df = pd.read_excel(_find_file(base, "Final Unit Setup View.xlsx"), header=8)
    df["unit_code"] = df["Unit number"].apply(clean_unit)
    return df

def load_rentable_items(base):
    df_raw = pd.read_excel(_find_file(base, "Rentable and Assignable Items Status.xls"),
                           engine=("xlrd" if _find_file(base, "Rentable and Assignable Items Status.xls").endswith(".xls") else None), header=None)
    records = []
    for _, row in df_raw.iterrows():
        vals = row.dropna().tolist()
        if not vals or not isinstance(vals[0], str): continue
        # Detect any row whose first non-null value starts with a rentable type name
        item_name = vals[0]
        if not re.match(r"[A-Za-z].+\s\d+$", item_name): continue  # "Garage 1", "Parking 3" etc
        item_num  = re.search(r"\d+$", item_name)
        item_num  = item_num.group().zfill(2) if item_num else "00"
        status = ""
        unit_code = None
        begin_date = None
        amount = 0.0
        for v in vals[1:4]:
            if isinstance(v, str) and v in ("In Use","Leased","NTV","Unassigned","Available"):
                status = v; break
        for v in vals:
            if isinstance(v, (int, float)) and 100 < v < 9999 and unit_code is None:
                unit_code = clean_unit(int(v))
            elif isinstance(v, str) and re.match(r"^\d{3,4}$", v.strip()):
                unit_code = clean_unit(int(v.strip()))
        for v in vals:
            if isinstance(v, (datetime, date)):
                begin_date = fmt_date(v); break
            if isinstance(v, str):
                try: begin_date = pd.to_datetime(v, format="%m/%d/%Y").strftime("%Y-%m-%d"); break
                except: pass
        for v in reversed(vals):
            if isinstance(v, float) and 10 <= v <= 1000:
                amount = v; break
        records.append({"item_name": item_name, "item_number": item_num,
                         "status": status, "unit_code": unit_code,
                         "begin_date": begin_date, "amount": amount})
    return pd.DataFrame(records) if records else pd.DataFrame(
        columns=["item_name","item_number","status","unit_code","begin_date","amount"])

def load_insurance(base):
    df = pd.read_excel(_find_file(base, "Renters Insurance Status Report.xlsx"), header=8)
    df = df[df["Unit"].notna()]
    df["unit_code"] = df["Unit"].apply(clean_unit)
    return df

def load_prospects(base):
    df = pd.read_excel(_find_file(base, "Prospect Contact Level Details .xlsx"), header=8)
    return df.drop_duplicates("Guest Card ID")

def load_all_unit(base):
    df = pd.read_excel(_find_file(base, "Final All Unit.xls"),
                       engine=("xlrd" if _find_file(base, "Final All Unit.xls").endswith(".xls") else None), header=6)
    df = df[df["Bldg/Unit"].notna()]
    df = df[df["Bldg/Unit"].apply(lambda x: bool(re.match(r"^\d", str(x))))]
    df["unit_code"] = df["Bldg/Unit"].apply(clean_unit)
    return df.drop_duplicates("unit_code")

def load_birthdays(base):
    df = pd.read_excel(_find_file(base, "Resident Birthdays.xls"),
                       engine=("xlrd" if _find_file(base, "Resident Birthdays.xls").endswith(".xls") else None), header=7)
    df = df[df["Name"].notna() & df["Unit"].notna()]
    df = df[df["Unit"].apply(lambda x: str(x).replace(".0","").isdigit() if pd.notna(x) else False)]
    df["unit_code"] = df["Unit"].apply(clean_unit)
    by_unit = {}
    for _, row in df.iterrows():
        uc = row["unit_code"]
        email = row.get("Email Address", "")
        phone = row.get("Home Phone", "")
        if uc and pd.notna(email) and "noemail" not in str(email).lower():
            by_unit.setdefault(uc, {})
            if "email" not in by_unit[uc]: by_unit[uc]["email"] = str(email)
        if uc and pd.notna(phone) and str(phone).strip():
            by_unit.setdefault(uc, {})
            if "phone" not in by_unit[uc]: by_unit[uc]["phone"] = clean_phone(phone)
    return by_unit


# ─────────────────────── TENANT BASE BUILDER ─────────────────────────────────

def build_tenant_base(rr, ld_idx, cld_by_unit, bday_by_unit, mappings, property_code):
    tenants = {}
    for _, rr_row in rr.iterrows():
        resh_id   = rr_row["resh_id"]
        unit_code = rr_row["unit_code"]
        if not resh_id or not unit_code: continue
        fp         = str(rr_row.get("Floorplan", "")).strip()
        status_raw = str(rr_row.get("Unit/Lease Status", "")).strip()
        status     = STATUS_MAP.get(status_raw, 0)
        last, first, middle = parse_name(str(rr_row.get("Name", "")) if pd.notna(rr_row.get("Name")) else "")
        ld  = ld_idx.loc[resh_id] if resh_id in ld_idx.index else None
        phone1 = phone2 = phone3 = None
        lease_sign = notice_date = move_out_sched = None
        lease_term = 12
        address1 = city = state = zipcode = None

        if ld is not None:
            phone1      = clean_phone(ld.get("Cell Phone"))
            phone2      = clean_phone(ld.get("Home Phone"))
            phone3      = clean_phone(ld.get("Work Phone"))
            lease_sign  = fmt_date(ld.get("Lease signed date"))
            notice_date = fmt_date(ld.get("Notice given date"))
            move_out_sched = fmt_date(ld.get("Scheduled move out"))
            lease_term  = extract_term(ld.get("Lease term")) or 12
            for af in ("FAS Forwarding address", "Billing address"):
                addr = ld.get(af, "")
                if pd.notna(addr) and str(addr).strip():
                    m = re.match(r"(.+?)\s*,\s*(.+?)\s*,\s*([A-Z]{2})\s+([\d-]+)", str(addr))
                    if m:
                        address1 = m.group(1).strip(); city = m.group(2).strip()
                        state    = m.group(3).strip(); zipcode = m.group(4).strip()
                        break

        if status == 4 and not notice_date:
            notice_date = fmt_date(rr_row.get("Move-Out"))

        email = None
        for cld_row in cld_by_unit.get(unit_code, []):
            if last and str(cld_row.get("Last Name", "")).lower() == last.lower():
                raw = cld_row.get("E-mail", "")
                if pd.notna(raw) and "noemail" not in str(raw).lower():
                    email = str(raw)
                if not phone1: phone1 = clean_phone(cld_row.get("Cell Phone"))
                if not phone2: phone2 = clean_phone(cld_row.get("Home Phone"))
                if not address1:
                    addr = cld_row.get("Address", "")
                    if pd.notna(addr) and str(addr).strip():
                        m = re.match(r"(.+?)\s*,\s*(.+?)\s*,\s*([A-Z]{2})\s+([\d-]+)", str(addr))
                        if m:
                            address1 = m.group(1).strip(); city = m.group(2).strip()
                            state    = m.group(3).strip(); zipcode = m.group(4).strip()
                break

        if not email:
            bday = bday_by_unit.get(unit_code, {})
            email = bday.get("email")
            if not phone1: phone1 = bday.get("phone")

        if not address1 and unit_code:
            prop_addr = mappings.get("address", "")
            address1  = f"{prop_addr} #{int(unit_code)}" if prop_addr else f"Unit #{int(unit_code)}"
            city      = mappings.get("city", "")
            state     = mappings.get("state", "")
            zipcode   = mappings.get("zipcode", "")

        tenants[resh_id] = {
            "resh_id": resh_id, "tenant_code": make_tenant_code(resh_id),
            "unit_code": unit_code, "status": status,
            "last_name": last, "first_name": first, "middle_name": middle or None,
            "dob": None, "ssn": None,
            "move_in":   fmt_date(rr_row.get("Move-In")),
            "move_out":  fmt_date(rr_row.get("Move-Out")) or move_out_sched,
            "notice_date": notice_date,
            "lease_from": fmt_date(rr_row.get("Lease Start")),
            "lease_to":   fmt_date(rr_row.get("Lease End")),
            "lease_sign": lease_sign,
            "phone1": phone1, "phone2": phone2, "phone3": phone3, "phone4": None,
            "email": email,
            "address1": address1, "address2": None,
            "city": city, "state": state, "zipcode": zipcode,
            "rent":    safe_int(rr_row.get("Lease Rent")),
            "deposit": safe_int(rr_row.get("Required\nDeposit")),
            "lease_term": lease_term, "floorplan": fp,
        }
    return tenants


# ─────────────────────── ETL GENERATORS ──────────────────────────────────────

TENANT_COLS = ["Property_Code","Tenant_Code","Unit_Code","Status","Last_Name",
               "First_Name","Middle_Name","Date_of_Birth","Social_Security_Number",
               "Move_In_Date","Move_Out_Date","Notice_Date","Lease_From_Date",
               "Lease_To_Date","Lease_Sign_Date","Phone_Number_1","Phone_Number_2",
               "Phone_Number_3","Phone_Number_4","Email","Address1","Address2",
               "City","State","Zipcode","Rent","Security_Deposit_0","LeaseTerm"]

def t_row(t, prop_code):
    return [prop_code, t["tenant_code"], t["unit_code"], t["status"],
            t["last_name"], t["first_name"], t["middle_name"], t["dob"], t["ssn"],
            t["move_in"], t["move_out"], t["notice_date"],
            t["lease_from"], t["lease_to"], t["lease_sign"],
            t["phone1"], t["phone2"], t["phone3"], t["phone4"], t["email"],
            t["address1"], t["address2"], t["city"], t["state"], t["zipcode"],
            t["rent"], t["deposit"], t["lease_term"]]

def gen_tenants(tenants, status_filter, prop_code, include_tcode=False):
    """include_tcode=False → Tenant_Code column is blank (Phase 1)."""
    rows = []
    for t in tenants.values():
        if t["status"] != status_filter:
            continue
        row = t_row(t, prop_code)
        if not include_tcode:
            row[1] = None   # blank Tenant_Code (index 1 in TENANT_COLS)
        rows.append(row)
    return TENANT_COLS, rows


def gen_roommates(all_res, tenants, prop_code):
    cols = ["Tenant_Code","Roommate_PhoneNumber1","Roommate_PhoneNumber2",
            "Roommate_PhoneNumber3","Roommate_PhoneNumber4","Roommate_Email",
            "Roommate_ALTEmail","Date_of_Birth","Roommate_SSN","Property_Code",
            "Unit_Code","Roommate_Salutation","Roommate_LastName","Roommate_FirstName",
            "Roommate_MoveIn","Roommate_MoveOut","Occupant_Type","Roommate_Occupant",
            "Roommate_ACHOptOut","Roommate_Relationship","Roommate_Notes","Roommate_Code"]
    unit_to_t = {t["unit_code"]: t for t in tenants.values()}
    curr = all_res[all_res["Status"] == "Current resident"]
    rows = []; counter = {}
    for uc, grp in curr.groupby("unit_code"):
        primary = unit_to_t.get(uc)
        if not primary: continue
        tc = primary["tenant_code"]; prim_last = primary["last_name"].lower()
        for idx, (_, person) in enumerate(grp.iterrows()):
            last, first, _ = parse_name(str(person["Name"]))
            if last.lower() == prim_last and idx == 0: continue
            phone = clean_phone(str(person.get("Phone", "")))
            counter[uc] = counter.get(uc, 0) + 1
            rcode = f"rr{uc}{counter[uc]:02d}"
            rows.append([tc, phone, None, None, None, None, None, None, None,
                         prop_code, uc, None, last, first,
                         fmt_date(person.get("Move In")), fmt_date(person.get("Move Out")),
                         1, 0, 0, "Roommate", None, rcode])
    return cols, rows

def gen_rentable_item_types(mappings, prop_code):
    cols = ["Property_Code","Charge_Code","RentableItemType_Code","Description","Rent","Taxable","Service_Charge"]
    rows = []
    for rt in mappings["rentable_types"]:
        rows.append([prop_code, rt["charge_code"], rt["type_name"].lower(), rt["type_name"], rt["rent"], 0, 0])
    if not rows:
        rows = [[prop_code, "garagmsc", "garage", "Garage", 95, 0, 0]]
    return cols, rows

def gen_rentable_items(rent_df, mappings, prop_code):
    cols = ["Property_Code","RentableItemType_Code","RentableItem_Code","Description","Rent"]
    rows = []; seen = set()
    # Determine type name from mappings
    type_name = mappings["rentable_types"][0]["type_name"] if mappings["rentable_types"] else "Garage"
    default_rent = mappings["rentable_types"][0]["rent"] if mappings["rentable_types"] else 95
    for _, row in rent_df.iterrows():
        num = str(row["item_number"])
        if num not in seen:
            seen.add(num)
            rows.append([prop_code, type_name, num, type_name, int(row.get("amount", default_rent))])
    rows.sort(key=lambda x: int(x[2]) if str(x[2]).isdigit() else 99)
    return cols, rows

def gen_ri_policies(ins_df, tenants, prop_code):
    cols = ["Policy_Number","Insurer_Name","Liability_Amount","Effective_Date","Expired_Date",
            "Cancel_Date","Tenant_Code","Unit_Number","First_Name","Last_Name","Tenant_Code1"]
    unit_to_t = {t["unit_code"]: t for t in tenants.values()}
    rows = []
    for _, row in ins_df.iterrows():
        t  = unit_to_t.get(row["unit_code"])
        tc = t["tenant_code"] if t else None
        policy  = str(row.get("Policy #", ""))  if pd.notna(row.get("Policy #"))  else None
        carrier = str(row.get("Carrier", ""))   if pd.notna(row.get("Carrier"))   else None
        rows.append([policy, carrier, 100000,
                     fmt_date(row.get("Policy Start Date")), fmt_date(row.get("Expiration Date")),
                     None, tc, 0, None, None, tc])
    return cols, rows

def gen_lease_charges(rr, tenants, prop_code):
    cols = ["Property_Code","Tenant_Code","From_Date","To_Date","Charge_Code",
            "Amount","Is_Hold","Is_Taxable","Is_Ach","Max_Month","Is_CreditCard",
            "RentableItemType_Code","RentableItem_Code","LateFee"]
    resh_to_t = {t["resh_id"]: t for t in tenants.values()}
    rows = []
    for _, rr_row in rr.iterrows():
        rid = rr_row.get("resh_id")
        if not rid or rid not in resh_to_t: continue
        t = resh_to_t[rid]
        for col, yardi_code in CHARGE_CODE_MAP.items():
            amt = rr_row.get(col, 0)
            if pd.isna(amt) or amt == 0: continue
            ri_type = "Garage" if col == "GARAGE" else None
            rows.append([prop_code, t["tenant_code"], t["lease_from"], None,
                         yardi_code, float(amt), 0, 0, 0, 0, 0, ri_type, None, 0])
    return cols, rows

def gen_manage_rentable(rent_df, tenants, mappings, prop_code):
    cols = ["Property_Code","Tenant_Code","RentableItemType_Code","RentableItem_Code","Lease_From"]
    unit_to_t = {t["unit_code"]: t for t in tenants.values()}
    type_name = mappings["rentable_types"][0]["type_name"] if mappings["rentable_types"] else "Garage"
    rows = []
    for _, row in rent_df.iterrows():
        if row["status"] not in ("In Use", "Leased", "NTV"): continue
        uc = row["unit_code"]
        if not uc or uc not in unit_to_t: continue
        t = unit_to_t[uc]
        rows.append([prop_code, t["tenant_code"], type_name, row["item_number"],
                     row.get("begin_date") or t["lease_from"]])
    return cols, rows

def gen_leasebut_demo(rr, ld_idx, cld_by_unit, tenants, prop_code):
    cols = ["TableName","HCODE","demo_tcode","Unnamed: 3","demo_name","Unnamed: 5",
            "SCODE","SNAME","demo_lease_role","demo_birthdate","demo_employer",
            "demo_job_title","demo_annual_salary","demo_emp_zip_code","demo_cur_res_type",
            "demo_move_reason","demo_prior_zip","demo_dist_work","demo_cnt_vehicles",
            "demo_cnt_pet","demo_gender","demo_marital_status"]
    resh_to_t = {t["resh_id"]: t for t in tenants.values()}
    rows = []
    for _, rr_row in rr.iterrows():
        rid = rr_row.get("resh_id")
        if not rid or rid not in resh_to_t: continue
        t  = resh_to_t[rid]
        uc = t["unit_code"]
        gender = marital = employer = job = income = None
        for cld in cld_by_unit.get(uc, []):
            if str(cld.get("Last Name", "")).lower() == t["last_name"].lower():
                gender   = cld.get("Gender")                             if pd.notna(cld.get("Gender", None))                             else None
                marital  = cld.get("Marital Status")                     if pd.notna(cld.get("Marital Status", None))                     else None
                employer = cld.get("Current Employment Name")            if pd.notna(cld.get("Current Employment Name", None))            else None
                job      = cld.get("Current Employment Job Title")       if pd.notna(cld.get("Current Employment Job Title", None))       else None
                try:
                    raw_inc = cld.get("Current Employment Estimated Annual Income")
                    income  = float(raw_inc) if pd.notna(raw_inc) else None
                except: income = None
                break
        rows.append(["leasebut_demo", rid, t["tenant_code"], None,
                     f"{t['last_name']}, {t['first_name']}", None, None, None,
                     "Head of household", t.get("dob"), employer, job, income,
                     None, None, None, None, None, None, None, gender, marital])
    return cols, rows

def gen_prospects(pros_df, prop_code, unit_type_map):
    cols = ["Prospect_Code","LastName","FirstName","MiddleName",
            "Address1","Address2","City","State","ZipCode","HowLong",
            "HomePhone","OfficePhone","CellPhone","Fax","FedId","Email",
            "Date_Of_Birth","Income","RelationShip","Occupant","LeaseStep",
            "Status","Preferred_Rent","Preferred_Bedrooms","Preferred_Bath",
            "Preferred_MoveIn","Property_Code","LeaseFrom","LeaseTo",
            "Source","Agent","First_Contacted_On","Notes","LeaseTerm","FirstContactType"]
    smap = {"Active": 6, "Lost": 7, "Unqualified": 8}
    lstep = {"Active": 10, "Lost": 7, "Unqualified": 8}
    rows = []
    for _, row in pros_df.iterrows():
        gid  = row.get("Guest Card ID")
        pcode = f"p{int(gid):07d}" if pd.notna(gid) else None
        last  = str(row.get("Last Name",""))  if pd.notna(row.get("Last Name"))  else None
        first = str(row.get("First Name","")) if pd.notna(row.get("First Name")) else None
        mid   = str(row.get("Middle Name",""))if pd.notna(row.get("Middle Name"))else None
        ph1   = clean_phone(row.get("1st Phone Number"))
        ph2   = clean_phone(row.get("2nd Phone Number"))
        email = str(row.get("E-mail","")) if pd.notna(row.get("E-mail")) else None
        stat  = str(row.get("Status", "Active"))
        fp    = str(row.get("Floor Plan","")) if pd.notna(row.get("Floor Plan")) else ""
        ut    = unit_type_map.get(fp, {})
        rd_raw = row.get("Price Desired", 0)
        try: rd = float(str(rd_raw).split("-")[0].strip().replace(",",""))
        except: rd = 0.0
        lt = extract_term(row.get("Lease Term"))
        src   = str(row.get("Primary advertising source","")) if pd.notna(row.get("Primary advertising source")) else None
        agent = str(row.get("Leasing Consultant",""))         if pd.notna(row.get("Leasing Consultant"))         else None
        notes = str(row.get("Notes",""))                      if pd.notna(row.get("Notes"))                      else None
        rows.append([pcode, last, first, mid, None, None, None, None, None, 0,
                     ph1, None, ph2, None, None, email, None, 0, None, 0,
                     lstep.get(stat, 10), smap.get(stat, 6), rd,
                     ut.get("beds", 1), ut.get("baths", 1),
                     fmt_date(row.get("Date Needed")), prop_code, None, None,
                     src, agent, fmt_date(row.get("Original guest card creation date")),
                     notes, lt or 12, "Other"])
    return cols, rows

def gen_unit_types(mappings, prop_code):
    cols = ["Property_Code","UnitType_Code","Description","Beds","Baths",
            "SQFT","Rent","Deposit","MinimumRent","Maximim_Rent"]
    rows = []
    for prior, ut in mappings["unit_type_map"].items():
        rows.append([prop_code, ut["yardi_code"], ut.get("desc",""),
                     ut["beds"], ut["baths"], ut["sqft"],
                     ut["rent"], 0, 0, 0])
    return cols, rows

def gen_comm_units(all_unit_df, unit_setup_df, mappings, prop_code):
    cols = ["Property_Code","Unit_Code","Unnamed: 2","Unit_Type",
            "Rental_Type","Country","Available_Date","Date_Ready",
            "Address_1","City","State","Zip_Code","Rent","SQFT",
            "BedRooms","Address_2","Exclude","Bldg_Code"]
    fp_by_unit = {}
    for _, row in unit_setup_df.iterrows():
        uc = row["unit_code"]
        fp_raw = str(row.get("Floor plan","")).split(" - ")[0].strip()
        if uc and fp_raw and fp_raw != "nan": fp_by_unit[uc] = fp_raw
    rows = []; seen = set()
    prop_addr = mappings.get("address","")
    city      = mappings.get("city","")
    state     = mappings.get("state","")
    zipcode   = mappings.get("zipcode","")
    for _, row in all_unit_df.iterrows():
        uc = row["unit_code"]
        if not uc or uc in seen: continue
        seen.add(uc)
        fp = fp_by_unit.get(uc, str(row.get("Floor plan","")).strip())
        ut = mappings["unit_type_map"].get(fp, {})
        yardi_fp    = ut.get("yardi_code", f"{mappings['prop_prefix']}0000")
        yardi_code2 = f"{prop_code}-{uc}"
        sqft = safe_int(str(row.get("SQFT","0")).replace(",",""), ut.get("sqft",0))
        rent = safe_int(str(row.get("Market Rent","0")).replace(",",""), ut.get("rent",0))
        addr = f"{prop_addr} #{int(uc)}" if prop_addr and uc.isdigit() else None
        rows.append([prop_code, uc, yardi_code2, yardi_fp,
                     "Residential","US", None, None,
                     addr, city, state, zipcode,
                     rent, sqft, ut.get("beds",1), None, 0, None])
    return cols, rows

def gen_property_amenities(mappings, prop_code):
    cols = ["Property_Code","Amenity_Code","Description",
            "Current_Charge","Current_Charge_Date",
            "Prior_Charge","Prior_Charge_Date",
            "Proposed_Charge","Proposed_Charge_Date"]
    asof = datetime.now().strftime("%Y-%m-%d")
    seen = set(); rows = []
    for prior, (desc, code, amt) in mappings["amenity_map"].items():
        if code not in seen:
            seen.add(code)
            rows.append([prop_code, code, desc, amt, asof, amt, asof, amt, asof])
    return cols, rows

def gen_unit_amenities(unit_setup_df, mappings, prop_code):
    cols = ["Property_Code","Unit_Code","Amenity_Name","Amenity_Code",
            "Amenity_Description","Current_Charge","Current_Charge_Date",
            "Prior_Charge","Prior_Charge_Date","Proposed_Charge","Proposed_Charge_Date"]
    asof = datetime.now().strftime("%Y-%m-%d")
    rows = []; seen_rent = set()
    for _, row in unit_setup_df.iterrows():
        uc = row["unit_code"]
        if not uc: continue
        if uc not in seen_rent:
            seen_rent.add(uc)
            rows.append([prop_code, uc, "Rent","Rent","Rent", 0,asof, 0,asof, 0,asof])
        aname = str(row.get("Unit amenity Name","")).strip()
        aval  = row.get("Unit amenity dollar value", 0)
        if aname and aname != "nan":
            mapped = mappings["amenity_map"].get(aname)
            if mapped:
                code, desc, default_amt = mapped[1], mapped[0], mapped[2]
            else:
                code  = re.sub(r"[^A-Za-z0-9_]","",aname)[:20]
                desc  = aname; default_amt = 0
            amt = float(aval) if pd.notna(aval) else default_amt
            rows.append([prop_code, uc, code, code, desc, amt,asof, amt,asof, amt,asof])
    return cols, rows


# ─────────────────────────── MAIN RUNNER ─────────────────────────────────────


def _apply_tcode(t, tcode_map):
    """Return the real Yardi tcode for this tenant, or None if not yet mapped."""
    return tcode_map.get(t["resh_id"]) or tcode_map.get(t["unit_code"])


def gen_roommates_p2(all_res, tenants, prop_code, tcode_map):
    """Phase 2: roommates with real tcodes from Yardi."""
    tc_to_rid  = {t["tenant_code"]: t["resh_id"] for t in tenants.values()}
    cols, rows = gen_roommates(all_res, tenants, prop_code)
    for row in rows:
        rid    = tc_to_rid.get(row[0])
        row[0] = tcode_map.get(rid) or row[0]
    return cols, rows


def gen_ri_policies_p2(ins_df, tenants, prop_code, tcode_map):
    """Phase 2: RI policies with real tcodes."""
    # Build reverse map: generated_tcode -> resh_id for lookup
    tc_to_rid = {t["tenant_code"]: t["resh_id"] for t in tenants.values()}
    cols, rows = gen_ri_policies(ins_df, tenants, prop_code)
    for row in rows:
        rid     = tc_to_rid.get(row[6])
        real_tc = tcode_map.get(rid) if rid else None
        row[6]  = real_tc or row[6]
        row[10] = real_tc or row[10]
    return cols, rows


def gen_lease_charges_p2(rr, tenants, prop_code, tcode_map):
    """Phase 2: lease charges with real tcodes."""
    tc_to_rid = {t["tenant_code"]: t["resh_id"] for t in tenants.values()}
    cols, rows = gen_lease_charges(rr, tenants, prop_code)
    for row in rows:
        rid     = tc_to_rid.get(row[1])
        row[1]  = tcode_map.get(rid) or row[1]
    return cols, rows


def gen_manage_rentable_p2(rent_df, tenants, mappings, prop_code, tcode_map):
    """Phase 2: rentable item assignments with real tcodes."""
    tc_to_rid = {t["tenant_code"]: t["resh_id"] for t in tenants.values()}
    cols, rows = gen_manage_rentable(rent_df, tenants, mappings, prop_code)
    for row in rows:
        rid    = tc_to_rid.get(row[1])
        row[1] = tcode_map.get(rid) or row[1]
    return cols, rows


def gen_leasebut_demo_p2(rr, ld_idx, cld_by_unit, tenants, prop_code, tcode_map):
    """Phase 2: demo file with real tcodes (demo_tcode column = col index 2)."""
    cols, rows = gen_leasebut_demo(rr, ld_idx, cld_by_unit, tenants, prop_code)
    for row in rows:
        rid     = row[1]   # HCODE = resh_id
        real_tc = tcode_map.get(rid) or row[2]
        row[2]  = real_tc
    return cols, rows


def run_conversion(base, output_dir, mappings, property_code, progress_cb=None, include_former_bal=True):
    """
    Run full conversion.
    base: path ending with '/' containing the OneSite export files
    mappings: result of load_takeover_guide()
    property_code: Yardi numeric property code string (e.g. "13400")
    """
    os.makedirs(output_dir, exist_ok=True)

    def log(msg):
        if progress_cb: progress_cb(msg)
        else: print(msg)

    log("📂 Loading source files...")
    rr         = load_rent_roll(base)
    ld_idx     = load_lease_details(base)
    cld        = load_contract_details(base)
    all_res    = load_all_residents(base)
    unit_setup = load_unit_setup(base)
    rent_items = load_rentable_items(base)
    insurance  = load_insurance(base)
    prospects  = load_prospects(base)
    all_unit   = load_all_unit(base)
    bdays      = load_birthdays(base)

    log("🔧 Building tenant records...")
    tenants = build_tenant_base(rr, ld_idx, cld, bdays, mappings, property_code)
    curr   = sum(1 for t in tenants.values() if t["status"] == 0)
    notice = sum(1 for t in tenants.values() if t["status"] == 4)
    future = sum(1 for t in tenants.values() if t["status"] == 6)
    log(f"   → {curr} current  |  {notice} on notice  |  {future} future")

    log("📋 Loading former residents with balance / eviction...")
    former_tenants = load_former_records(base)
    evictions  = sum(1 for t in former_tenants.values() if t["status"] == 10)
    former_bal = sum(1 for t in former_tenants.values() if t["status"] == 5)
    # Also flag current residents who have eviction proceedings
    try:
        ld_evict = pd.read_excel(_find_file(base, "Lease Details .xlsx"), header=8)
        evict_ids_curr = set(
            ld_evict[ld_evict["Eviction proceedings started"] == "Yes"]["Household ID/ Resh ID"]
            .dropna().apply(lambda x: safe_int(x, 0))
        )
        for rid, t in tenants.items():
            if rid in evict_ids_curr and t["status"] == 0:
                t["status"] = 10   # promote current resident to eviction
    except Exception:
        pass
    # Merge former records (don't overwrite existing current/notice/future tenants)
    for rid, t in former_tenants.items():
        if rid not in tenants:
            tenants[rid] = t
    log(f"   → {evictions} evictions  |  {former_bal} former with balance")

    prop_code = property_code
    outputs = [
        ("ETL_ResTenants",             *gen_tenants(tenants, 0, prop_code)),
        ("ETL_ResTenants_Eviction",    *gen_tenants(tenants, 10, prop_code)),
        ("ETL_ResTenants_Notice",      *gen_tenants(tenants, 4, prop_code)),
        ("ETL_ResTenants_Future",      *gen_tenants(tenants, 6, prop_code)),
    ]
    if include_former_bal:
        outputs.append(("ETL_ResTenants_FormerBal", *gen_tenants(tenants, 5, prop_code)))
    outputs += [
        ("ETL_ResRoommates",          *gen_roommates(all_res, tenants, prop_code)),
        ("ETL_ResRentableItemsTypes", *gen_rentable_item_types(mappings, prop_code)),
        ("ETL_ResRentableItems",      *gen_rentable_items(rent_items, mappings, prop_code)),
        ("ETL_RIPolicies",            *gen_ri_policies(insurance, tenants, prop_code)),
        ("ETL_ResLeaseCharges",       *gen_lease_charges(rr, tenants, prop_code)),
        ("ETL_ResManageRentableItems",*gen_manage_rentable(rent_items, tenants, mappings, prop_code)),
        ("ETL_leasebut_demo",         *gen_leasebut_demo(rr, ld_idx, cld, tenants, prop_code)),
        ("ETL_ResProspects",          *gen_prospects(prospects, prop_code, mappings["unit_type_map"])),
        ("ETL_ResUnitTypes",          *gen_unit_types(mappings, prop_code)),
        ("ETL_CommUnits",             *gen_comm_units(all_unit, unit_setup, mappings, prop_code)),
        ("ETL_ResPropertyAmenities",  *gen_property_amenities(mappings, prop_code)),
        ("ETL_ResUnitAmenities",      *gen_unit_amenities(unit_setup, mappings, prop_code)),
    ]

    ts = datetime.now().strftime("%y%m%d_%H%M%S")
    generated = []
    for table_name, cols, rows in outputs:
        fname = f"{ts}_{table_name}.xlsx"
        fpath = os.path.join(output_dir, fname)
        write_etl_xlsx(table_name, cols, rows, fpath)
        generated.append(fpath)
        log(f"   ✅ {table_name} → {len(rows)} rows")

    zip_path = os.path.join(output_dir, f"{ts}_ETL_Output.zip")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for fp in generated:
            zf.write(fp, os.path.basename(fp))

    log(f"\n🎉 Done!  {len(generated)} files  →  {zip_path}")
    return generated, zip_path


# ─────────────────── PHASE 1: RESIDENTS + PROPERTY FILES ────────────────────

PHASE1_FILES = {
    "ETL_ResTenants", "ETL_ResTenants_Notice", "ETL_ResTenants_Future",
    "ETL_ResTenants_Eviction", "ETL_ResTenants_FormerBal",
    "ETL_ResRentableItemsTypes", "ETL_ResRentableItems",
    "ETL_ResUnitTypes", "ETL_CommUnits",
    "ETL_ResPropertyAmenities", "ETL_ResUnitAmenities", "ETL_ResProspects",
}

PHASE2_FILES = {
    "ETL_ResRoommates", "ETL_RIPolicies", "ETL_ResLeaseCharges",
    "ETL_ResManageRentableItems", "ETL_leasebut_demo",
}


def run_phase1(base, output_dir, mappings, property_code,
               include_former_bal=True, progress_cb=None):
    """
    Phase 1: Generate resident files (Tenant_Code BLANK) + all property files.
    Returns (generated_files, zip_path, tenants).
    """
    os.makedirs(output_dir, exist_ok=True)

    def log(m):
        if progress_cb: progress_cb(m)
        else: print(m)

    log("📂 Loading source files...")
    rr         = load_rent_roll(base)
    ld_idx     = load_lease_details(base)
    cld        = load_contract_details(base)
    all_res    = load_all_residents(base)
    unit_setup = load_unit_setup(base)
    rent_items = load_rentable_items(base)
    insurance  = load_insurance(base)
    prospects  = load_prospects(base)
    all_unit   = load_all_unit(base)
    bdays      = load_birthdays(base)

    log("🔧 Building tenant records...")
    tenants = build_tenant_base(rr, ld_idx, cld, bdays, mappings, property_code)
    # Promote evictions + add former-with-balance
    try:
        ld_ev = pd.read_excel(_find_file(base, "Lease Details .xlsx"), header=8)
        evict_ids = set(
            ld_ev[ld_ev["Eviction proceedings started"] == "Yes"]
            ["Household ID/ Resh ID"].dropna()
            .apply(lambda x: safe_int(x, 0))
        )
        for rid, t in tenants.items():
            if rid in evict_ids and t["status"] == 0:
                t["status"] = 10
    except Exception:
        pass
    former_recs = load_former_records(base)
    for rid, t in former_recs.items():
        if rid not in tenants:
            tenants[rid] = t

    curr   = sum(1 for t in tenants.values() if t["status"] == 0)
    notice = sum(1 for t in tenants.values() if t["status"] == 4)
    future = sum(1 for t in tenants.values() if t["status"] == 6)
    evict  = sum(1 for t in tenants.values() if t["status"] == 10)
    fbal   = sum(1 for t in tenants.values() if t["status"] == 5)
    log(f"   → {curr} current | {notice} notice | {future} future | {evict} eviction | {fbal} former/bal")

    prop_code = property_code

    # Resident files — Tenant_Code column is BLANK (include_tcode=False)
    outputs = [
        ("ETL_ResTenants",         *gen_tenants(tenants, 0,  prop_code, include_tcode=False)),
        ("ETL_ResTenants_Notice",  *gen_tenants(tenants, 4,  prop_code, include_tcode=False)),
        ("ETL_ResTenants_Future",  *gen_tenants(tenants, 6,  prop_code, include_tcode=False)),
        ("ETL_ResTenants_Eviction",*gen_tenants(tenants, 10, prop_code, include_tcode=False)),
    ]
    if include_former_bal:
        outputs.append(
            ("ETL_ResTenants_FormerBal", *gen_tenants(tenants, 5, prop_code, include_tcode=False))
        )
    # Property files — no tcode needed
    outputs += [
        ("ETL_ResRentableItemsTypes", *gen_rentable_item_types(mappings, prop_code)),
        ("ETL_ResRentableItems",      *gen_rentable_items(rent_items, mappings, prop_code)),
        ("ETL_ResUnitTypes",          *gen_unit_types(mappings, prop_code)),
        ("ETL_CommUnits",             *gen_comm_units(all_unit, unit_setup, mappings, prop_code)),
        ("ETL_ResPropertyAmenities",  *gen_property_amenities(mappings, prop_code)),
        ("ETL_ResUnitAmenities",      *gen_unit_amenities(unit_setup, mappings, prop_code)),
        ("ETL_ResProspects",          *gen_prospects(prospects, prop_code, mappings["unit_type_map"])),
    ]

    ts = datetime.now().strftime("%y%m%d_%H%M%S")
    generated = []
    for table_name, cols, rows in outputs:
        fname = f"{ts}_{table_name}.xlsx"
        fpath = os.path.join(output_dir, fname)
        write_etl_xlsx(table_name, cols, rows, fpath)
        generated.append(fpath)
        log(f"   ✅ {table_name} → {len(rows)} rows")

    zip_path = os.path.join(output_dir, f"{ts}_Phase1_ETL.zip")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for fp in generated:
            zf.write(fp, os.path.join("Phase1_ETL", os.path.basename(fp)))

    log(f"\n🎉 Phase 1 complete — {len(generated)} files → {zip_path}")
    return generated, zip_path, tenants


# ─────────────────── PHASE 2: TCODE-DEPENDENT FILES ─────────────────────────

def load_tcode_mapping(etl_file_paths, tenants):
    """
    Read one or more ETL_ResTenants files exported from Yardi after Phase 1 import.
    Yardi populates the Tenant_Code column; this builds:
        resh_id (int) → yardi_tenant_code (str)
    matching by Unit_Code + Last_Name (handles Yardi writing unit as int, e.g. 117 not '0117').

    Pass a single path or a list of paths — all variant files are merged into one map.
    ETL format: row 1 = table name, row 2 = column headers, row 3+ = data.
    """
    if isinstance(etl_file_paths, (str, bytes, os.PathLike)):
        etl_file_paths = [etl_file_paths]

    # Build reverse lookup: (zero_padded_unit, last_name_lower) → resh_id
    key_to_rid = {}
    for rid, t in tenants.items():
        uc   = (t.get("unit_code") or "").strip()   # already zero-padded e.g. '0117'
        last = (t.get("last_name") or "").lower().strip()
        if uc and last:
            key_to_rid[(uc, last)] = rid

    tcode_map = {}
    for path in etl_file_paths:
        df = pd.read_excel(path, skiprows=1, header=0)
        for _, row in df.iterrows():
            # Normalize unit_code: Yardi may return int 117 — zero-pad to '0117'
            raw = row.get("Unit_Code", "") or ""
            s   = re.sub(r"['\s#]", "", str(raw))
            mm  = re.match(r"(\d+)", s)
            unit = mm.group(1).zfill(4) if mm else s.strip()

            last = str(row.get("Last_Name",   "") or "").strip().lower()
            tc   = str(row.get("Tenant_Code", "") or "").strip()

            if unit and last and tc and tc.lower() not in ("nan", "none", ""):
                rid = key_to_rid.get((unit, last))
                if rid and rid not in tcode_map:   # first match wins
                    tcode_map[rid] = tc

    return tcode_map



def run_phase2(base, output_dir, mappings, property_code,
               tenants, tcode_map, progress_cb=None):
    """
    Phase 2: Generate tcode-dependent files using the mapping returned by Yardi.
    tenants: the tenant dict from Phase 1 (stored in session state).
    tcode_map: {resh_id → yardi_tenant_code}
    """
    os.makedirs(output_dir, exist_ok=True)

    def log(m):
        if progress_cb: progress_cb(m)
        else: print(m)

    matched   = sum(1 for t in tenants.values() if tcode_map.get(t["resh_id"]))
    unmatched = sum(1 for t in tenants.values() if not tcode_map.get(t["resh_id"]))
    log(f"🔑 Tcode mapping: {matched} matched, {unmatched} unmatched")

    log("📂 Reloading dependent source files...")
    rr         = load_rent_roll(base)
    ld_idx     = load_lease_details(base)
    cld        = load_contract_details(base)
    all_res    = load_all_residents(base)
    rent_items = load_rentable_items(base)
    insurance  = load_insurance(base)

    prop_code = property_code
    outputs = [
        ("ETL_ResRoommates",          *gen_roommates_p2(all_res, tenants, prop_code, tcode_map)),
        ("ETL_RIPolicies",            *gen_ri_policies_p2(insurance, tenants, prop_code, tcode_map)),
        ("ETL_ResLeaseCharges",       *gen_lease_charges_p2(rr, tenants, prop_code, tcode_map)),
        ("ETL_ResManageRentableItems",*gen_manage_rentable_p2(rent_items, tenants, mappings, prop_code, tcode_map)),
        ("ETL_leasebut_demo",         *gen_leasebut_demo_p2(rr, ld_idx, cld, tenants, prop_code, tcode_map)),
    ]

    ts = datetime.now().strftime("%y%m%d_%H%M%S")
    generated = []
    for table_name, cols, rows in outputs:
        fname = f"{ts}_{table_name}.xlsx"
        fpath = os.path.join(output_dir, fname)
        write_etl_xlsx(table_name, cols, rows, fpath)
        generated.append(fpath)
        log(f"   ✅ {table_name} → {len(rows)} rows")

    zip_path = os.path.join(output_dir, f"{ts}_Phase2_ETL.zip")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for fp in generated:
            zf.write(fp, os.path.join("Phase2_ETL", os.path.basename(fp)))

    log(f"\n🎉 Phase 2 complete → {zip_path}")
    return generated, zip_path



# ─────────────────── VALIDATION WORKBOOK ─────────────────────────────────────

def build_validation_workbook(vdata, mappings, property_code, tenants, output_path):
    """
    Build a multi-sheet Excel validation workbook.
    Sheets:
      1. Summary       — counts by status, quality flags, property info
      2. Tenants        — full roster (all statuses, all columns)
      3. Quality Flags  — tenants missing email / phone / lease sign date
      4. Unit Types     — floor plan → Yardi code mapping
      5. Amenities      — amenity code + monthly charge mapping
      6. Charge Codes   — OneSite column → Yardi transaction code mapping
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.utils import get_column_letter
    from datetime import datetime as _dt

    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    # ── colour palette ────────────────────────────────────────────────────────
    DARK_BLUE   = "1F4E79"
    MID_BLUE    = "2E75B6"
    LIGHT_BLUE  = "BDD7EE"
    GREEN       = "375623"
    LIGHT_GREEN = "E2EFDA"
    ORANGE      = "C65911"
    LIGHT_ORANGE= "FCE4D6"
    RED         = "C00000"
    LIGHT_RED   = "FFDBE0"
    GREY_FILL   = "F2F2F2"
    WHITE       = "FFFFFF"
    YELLOW      = "FFFF00"

    def hdr_font(bold=True, colour=WHITE, sz=10):
        return Font(name="Calibri", bold=bold, color=colour, size=sz)

    def data_font(bold=False, colour="000000", sz=9):
        return Font(name="Calibri", bold=bold, color=colour, size=sz)

    def fill(hex_colour):
        return PatternFill("solid", start_color=hex_colour)

    def thin_border():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)

    def set_col_widths(ws, widths):
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    def write_header_row(ws, row_num, headers, bg=DARK_BLUE, fg=WHITE, height=20):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row_num, ci, h)
            c.font      = hdr_font(colour=fg)
            c.fill      = fill(bg)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row_num].height = height

    def write_data_row(ws, row_num, values, bg=None, bold=False):
        for ci, v in enumerate(values, 1):
            c = ws.cell(row_num, ci, v)
            c.font   = data_font(bold=bold)
            c.border = thin_border()
            if bg:
                c.fill = fill(bg)
            c.alignment = Alignment(vertical="center", wrap_text=True)

    # ═══════════════════════════════════════════════════════
    #  SHEET 1 — SUMMARY
    # ═══════════════════════════════════════════════════════
    ws1 = wb.create_sheet("Summary")
    S   = vdata["summary"]
    Q   = vdata["quality"]

    # Property header block
    ws1["A1"] = "RPM Living · Yardi ETL Conversion — Validation Report"
    ws1["A1"].font = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
    ws1["A2"] = f"Property: {mappings.get('prop_name','Unknown')}  |  Code: {property_code}  |  Generated: {_dt.now().strftime('%Y-%m-%d %H:%M')}"
    ws1["A2"].font = Font(name="Calibri", italic=True, size=10, color="595959")
    ws1.merge_cells("A1:H1"); ws1.merge_cells("A2:H2")
    ws1.row_dimensions[1].height = 24; ws1.row_dimensions[2].height = 16

    # Tenant counts section
    ws1["A4"] = "TENANT COUNTS"
    ws1["A4"].font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
    counts = [
        ("Current Residents",       S["current"],    MID_BLUE,    WHITE),
        ("On Notice",               S["notice"],     "C65911",    WHITE),
        ("Future / Applicants",     S["future"],     "375623",    WHITE),
        ("Eviction Proceedings",    S["eviction"],   "C00000",    WHITE),
        ("Former w/ Balance",       S["former_bal"], "7030A0",    WHITE),
        ("Total Tenant Records",    S["current"]+S["notice"]+S["future"]+S["eviction"]+S["former_bal"], DARK_BLUE, WHITE),
    ]
    write_header_row(ws1, 5, ["Category","Count"], bg=DARK_BLUE)
    for ri, (label, count, bg, fg) in enumerate(counts, 6):
        ws1.cell(ri, 1, label).font  = data_font(bold=(ri==11))
        ws1.cell(ri, 1).fill         = fill(LIGHT_BLUE if ri % 2 == 0 else WHITE)
        ws1.cell(ri, 1).border       = thin_border()
        ws1.cell(ri, 2, count).font  = data_font(bold=True)
        ws1.cell(ri, 2).alignment    = Alignment(horizontal="center")
        ws1.cell(ri, 2).border       = thin_border()

    # Property info section
    ws1["D4"] = "PROPERTY INFO"
    ws1["D4"].font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
    prop_info = [
        ("Property Name",  mappings.get("prop_name","")),
        ("Property Code",  property_code),
        ("Address",        mappings.get("address","")),
        ("City / State",   f"{mappings.get('city','')} {mappings.get('state','')} {mappings.get('zipcode','')}"),
        ("Prefix",         mappings.get("prop_prefix","")),
        ("Total Units",    S["total_units"]),
        ("Unit Types",     len(mappings.get("unit_type_map",{}))),
        ("Amenity Types",  len(mappings.get("amenity_map",{}))),
        ("RI Policies",    S["ri_policies"]),
        ("Prospects",      S["prospects"]),
    ]
    write_header_row(ws1, 5, ["","Field","Value"], bg=DARK_BLUE)
    for ri, (field, val) in enumerate(prop_info, 6):
        ws1.cell(ri, 4, field).font   = data_font(bold=True)
        ws1.cell(ri, 4).border        = thin_border()
        ws1.cell(ri, 4).fill          = fill(LIGHT_BLUE if ri%2==0 else WHITE)
        ws1.cell(ri, 5, str(val)).font= data_font()
        ws1.cell(ri, 5).border        = thin_border()
        ws1.cell(ri, 5).fill          = fill(LIGHT_BLUE if ri%2==0 else WHITE)

    # Quality flags section
    row = 17
    ws1.cell(row, 1, "DATA QUALITY FLAGS").font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
    row += 1
    write_header_row(ws1, row, ["Flag","Count","Status"], bg=MID_BLUE)
    row += 1
    quality_checks = [
        ("Missing Email",          len(Q["no_email"]),  len(Q["no_email"])==0),
        ("Missing Phone",          len(Q["no_phone"]),  len(Q["no_phone"])==0),
        ("Missing Lease Sign Date",len(Q["no_sign"]),   len(Q["no_sign"])==0),
        ("Unmapped Floor Plans",   len(vdata["unmapped_ut"]), len(vdata["unmapped_ut"])==0),
    ]
    for flag, count, ok in quality_checks:
        bg_row = LIGHT_GREEN if ok else LIGHT_RED
        ws1.cell(row, 1, flag).font   = data_font()
        ws1.cell(row, 1).fill         = fill(bg_row); ws1.cell(row, 1).border = thin_border()
        ws1.cell(row, 2, count).font  = data_font(bold=True)
        ws1.cell(row, 2).fill         = fill(bg_row); ws1.cell(row, 2).border = thin_border()
        ws1.cell(row, 2).alignment    = Alignment(horizontal="center")
        status_txt = "✓ OK" if ok else f"⚠ {count} issue(s)"
        ws1.cell(row, 3, status_txt).font  = data_font(bold=True, colour=GREEN if ok else RED)
        ws1.cell(row, 3).fill              = fill(bg_row); ws1.cell(row, 3).border = thin_border()
        row += 1

    set_col_widths(ws1, [28, 10, 2, 22, 32])
    ws1.freeze_panes = "A3"


    # ═══════════════════════════════════════════════════════
    #  SHEET 2 — FULL TENANT ROSTER
    # ═══════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Tenants")
    STATUS_LABEL = {0:"Current",4:"Notice",6:"Future",10:"Eviction",5:"Former/Balance"}
    STATUS_COLOUR= {0:LIGHT_BLUE, 4:LIGHT_ORANGE, 6:LIGHT_GREEN, 10:LIGHT_RED, 5:"EAD1DC"}

    hdrs2 = ["Unit","Status","Last Name","First Name","Move In","Lease From","Lease To",
             "Sign Date","Rent","Deposit","Term","Phone 1","Phone 2","Email","Address","City","State","Zip"]
    write_header_row(ws2, 1, hdrs2, bg=DARK_BLUE, height=22)

    all_tenants = sorted(tenants.values(), key=lambda t: (t.get("unit_code") or "9999", t.get("last_name","")))
    for ri, t in enumerate(all_tenants, 2):
        status_code = t.get("status", 0)
        row_bg = STATUS_COLOUR.get(status_code, WHITE)
        vals = [
            t.get("unit_code"), STATUS_LABEL.get(status_code, str(status_code)),
            t.get("last_name"), t.get("first_name"),
            t.get("move_in"),   t.get("lease_from"), t.get("lease_to"), t.get("lease_sign"),
            t.get("rent"),      t.get("deposit"),     t.get("lease_term"),
            t.get("phone1"),    t.get("phone2"),      t.get("email"),
            t.get("address1"),  t.get("city"),        t.get("state"),  t.get("zipcode"),
        ]
        write_data_row(ws2, ri, vals, bg=row_bg)

    set_col_widths(ws2, [7,12,16,14,11,11,11,11,9,9,6,14,14,28,30,14,6,10])
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:R1"


    # ═══════════════════════════════════════════════════════
    #  SHEET 3 — QUALITY FLAGS
    # ═══════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Quality Flags")
    ws3["A1"] = "Tenants with data quality issues — review before importing to Yardi"
    ws3["A1"].font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
    ws3.merge_cells("A1:H1")

    sections = [
        ("Missing Email",          Q["no_email"],  LIGHT_RED),
        ("Missing Phone Number",   Q["no_phone"],  LIGHT_ORANGE),
        ("Missing Lease Sign Date",Q["no_sign"],   LIGHT_BLUE),
    ]

    row = 3
    for section_title, items, bg in sections:
        if not items:
            continue
        ws3.cell(row, 1, f"{section_title} ({len(items)} tenant(s))").font = \
            Font(name="Calibri", bold=True, size=10, color=DARK_BLUE)
        ws3.merge_cells(f"A{row}:H{row}")
        row += 1
        write_header_row(ws3, row, ["Unit","Status","Last Name","First Name","Move In","Lease From","Rent","Email"], bg=MID_BLUE)
        row += 1
        for t in sorted(items, key=lambda x: x.get("unit_code") or ""):
            vals = [
                t.get("unit_code"),
                STATUS_LABEL.get(t.get("status",0),"?"),
                t.get("last_name"), t.get("first_name"),
                t.get("move_in"),   t.get("lease_from"),
                t.get("rent"),      t.get("email") or "MISSING",
            ]
            write_data_row(ws3, row, vals, bg=bg)
            row += 1
        row += 1   # blank row between sections

    set_col_widths(ws3, [7,12,16,14,11,11,9,30])
    ws3.freeze_panes = "A2"


    # ═══════════════════════════════════════════════════════
    #  SHEET 4 — UNIT TYPES
    # ═══════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Unit Types")
    hdrs4 = ["OneSite Code","Yardi Code","Description","Beds","Baths","SQFT","Market Rent","Total Units","Occupied","Status"]
    write_header_row(ws4, 1, hdrs4, bg=DARK_BLUE)
    for ri, r in enumerate(vdata["unit_types"], 2):
        bg = LIGHT_GREEN if r["Status"]=="✅ Mapped" else LIGHT_RED
        write_data_row(ws4, ri, [r[k] for k in hdrs4], bg=bg)
    set_col_widths(ws4, [14,13,24,6,7,8,12,11,9,12])
    ws4.freeze_panes = "A2"


    # ═══════════════════════════════════════════════════════
    #  SHEET 5 — AMENITIES
    # ═══════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Amenities")
    hdrs5 = ["OneSite Name","RPM Description","Yardi Code","Monthly Amt ($)","Units","Status"]
    write_header_row(ws5, 1, hdrs5, bg=DARK_BLUE)
    for ri, r in enumerate(vdata["amenities"], 2):
        bg = LIGHT_GREEN if "Mapped" in r["Status"] else LIGHT_ORANGE
        write_data_row(ws5, ri, [r[k] for k in hdrs5], bg=bg)
    set_col_widths(ws5, [30,28,18,14,7,12])
    ws5.freeze_panes = "A2"


    # ═══════════════════════════════════════════════════════
    #  SHEET 6 — CHARGE CODES
    # ═══════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Charge Codes")
    hdrs6 = ["OneSite Column","Yardi Code","Active Leases","Monthly Total ($)","Status"]
    write_header_row(ws6, 1, hdrs6, bg=DARK_BLUE)
    for ri, r in enumerate(vdata["charges"], 2):
        bg = LIGHT_GREEN if "Active" in r["Status"] else GREY_FILL
        write_data_row(ws6, ri, [r[k] for k in hdrs6], bg=bg)
    set_col_widths(ws6, [18,13,13,16,12])
    ws6.freeze_panes = "A2"

    # ── Finish ────────────────────────────────────────────────────────────────
    # Set Summary as the active sheet on open
    wb.active = ws1

    wb.save(output_path)
    return output_path
